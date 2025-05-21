import pandas as pd
from openpyxl import load_workbook
import os

def generate_team_sheets_from_schedule(
    xlsx_path,
    template_path,
    output_dir,
    target_sheet="Event",
    start_col=4,
    start_row=2
):
    """
    Reads each sheet from `xlsx_path` and fills the data into a Dynamics submission Excel template.
    One output file is created per team in the specified `output_dir`.

    Parameters:
    - xlsx_path: str - path to input file with multiple team sheets
    - template_path: str - path to Dynamics submission Excel template
    - output_dir: str - where to save team-specific output files
    - target_sheet: str - sheet name in the template to insert into
    - start_col: int - column index to begin writing data (default: 4, i.e., column D)
    - start_row: int - row index to begin writing data (default: 2)
    """

    # Ensure output folder exists
    os.makedirs(output_dir, exist_ok=True)

    # Load Excel file with team sheets
    xls = pd.ExcelFile(xlsx_path)

    for sheet_name in xls.sheet_names:
        df = xls.parse(sheet_name)

        # Load a fresh copy of the template each time
        wb = load_workbook(template_path)
        ws = wb[target_sheet]

        # Write data starting at column D (or specified start_col)
        for i, row in df.iterrows():
            for j, val in enumerate(row):
                ws.cell(row=start_row + i, column=start_col + j, value=val)

        # Save one file per team/sheet
        sanitized_name = sheet_name.replace(" ", "_")[:31]
        output_path = os.path.join(output_dir, f"{sanitized_name}.xlsx")
        wb.save(output_path)
        print(f"✅ Saved: {output_path}")
